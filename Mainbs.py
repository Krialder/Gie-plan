import tkinter as tk
from tkinter import messagebox
import random
import openpyxl
import json
import os
import datetime
import threading

# Create the main Tkinter window
root = tk.Tk()
root.title("Gießplan Generator")

# Define constants for configuration
MAX_WEEK = 52
DEFAULT_WEIGHTS = 1

# Add a configurable number of weeks for schedule generation
NUM_WEEKS = 6  # Default number of weeks, can be adjusted dynamically

# Encapsulate global variables into a class
class WateringData:
    def __init__(self):
        self.file_path = "people.json"  # Default file path
        self.default_file_path = "people.json"  # Configurable default file path
        self.yearly_file_template = "people_{year}.json"  # Template for yearly file paths
        self.people = []
        self.weights = []
        self.watering_history = {}
        self.data_modified = False  # Flag to track changes
        self.undo_stack = []  # Encapsulate undo stack
        self.load_data()

    # Enhanced JSON error handling
    def load_data(self):
        if os.path.exists(self.file_path):
            try:
                with open(self.file_path, "r") as file:
                    data = json.load(file)
                    self.people = data.get("PEOPLE", [])
                    self.weights = data.get("WEIGHTS", [])
                    self.watering_history = data.get("WATERING_HISTORY", {})

                    # Validate keys
                    if not isinstance(self.people, list) or not isinstance(self.weights, list) or not isinstance(self.watering_history, dict):
                        raise ValueError("Invalid JSON structure.")

            except json.JSONDecodeError:
                messagebox.showerror("Error", "Failed to decode JSON file. Initializing defaults.")
                self.initialize_defaults()
            except ValueError:
                messagebox.showerror("Error", "Invalid JSON structure. Initializing defaults.")
                self.initialize_defaults()
            except PermissionError:
                messagebox.showerror("Error", "Permission denied while accessing JSON file. Initializing defaults.")
                self.initialize_defaults()
        else:
            self.initialize_defaults()

    def initialize_defaults(self):
        self.people = ["Alice", "Bob", "Charlie", "Diana", "Eve", "Frank"]
        self.weights = [DEFAULT_WEIGHTS] * len(self.people)
        self.watering_history = {person: [] for person in self.people}
        self.file_path = self.default_file_path  # Use default file path
        self.save_data()

    def save_data(self):
        if self.data_modified:  # Only save if data has been modified
            with open(self.file_path, "w") as file:
                json.dump({"PEOPLE": self.people, "WEIGHTS": self.weights, "WATERING_HISTORY": self.watering_history}, file)
            self.data_modified = False  # Reset the flag

    def push_undo_action(self, action):
        """Push an action to the undo stack."""
        self.undo_stack.append(action)

    def pop_undo_action(self):
        """Pop an action from the undo stack."""
        return self.undo_stack.pop() if self.undo_stack else None

# Instantiate the data class
watering_data = WateringData()

# Ensure names in PEOPLE contain only letters
watering_data.people = [person for person in watering_data.people if person.isalpha()]

# Start with the current calendar week
current_week = datetime.date.today().isocalendar()[1]

# Function to find the next free calendar week with year transitions
def find_next_free_week():
    used_weeks = set()

    # Collect all used weeks from watering history
    for history in watering_data.watering_history.values():
        if isinstance(history, list):  # Ensure history is iterable
            used_weeks.update(
                int(entry.split()[1]) for entry in history if entry.startswith("Week") and entry.split()[1].isdigit()
            )

    # Find the next free week
    next_week = current_week
    while next_week in used_weeks:
        next_week += 1
        if next_week > MAX_WEEK:
            next_week = 1  # Transition to the next year

    return next_week

# Improved weight adjustment logic
def update_data():
    max_weight = 10  # Maximum weight for a person
    min_weight = DEFAULT_WEIGHTS  # Use DEFAULT_WEIGHTS as the minimum weight

    # Use dictionary for faster lookups
    watering_counts = {
        person: len(watering_data.watering_history.get(person, [])) for person in watering_data.people
    }

    for person in watering_data.people:
        watering_data.weights[watering_data.people.index(person)] = max(
            min_weight, max_weight - (watering_counts[person] * 2)  # Example: reduce weight more aggressively for frequent watering
        )

    watering_data.save_data()

# Combine update_statistics and update_weights into a single function
def generate_schedule():
    update_data()  # Update statistics before generating the schedule

    if len(watering_data.people) < 2:
        messagebox.showerror("Error", "Insufficient people to generate a schedule.")
        return []

    if all(weight <= 0 for weight in watering_data.weights):
        messagebox.showerror("Error", "Invalid weights configuration. Cannot generate schedule.")
        return []

    schedule = []
    selection_count = {person: len(watering_data.watering_history[person]) for person in watering_data.people}  # Track how often each person is selected

    # Get the current year
    current_year = datetime.date.today().year

    # Determine the starting week based on the last recorded week in the JSON file
    if any(entry.startswith("Week") for person in watering_data.watering_history.values() for entry in person):
        last_week = max(
            [int(entry.split()[1]) for person in watering_data.watering_history.values() for entry in person if entry.startswith("Week")]
        )
        start_week = last_week + 1
    else:
        start_week = current_week + 1

    week = start_week
    while len(schedule) < NUM_WEEKS:  # Use configurable number of weeks
        if week > MAX_WEEK:
            # Transition to the next year
            current_year += 1
            week = 1
            watering_data.watering_history["last_year"] = current_year

            # Create a new table in the Excel file for the new year
            save_to_excel([], watering_data.people, watering_data.watering_history, new_year=True)

        # Adjust weights based on selection intervals
        adjusted_weights = [watering_data.weights[i] / (1 + selection_count[watering_data.people[i]]) for i in range(len(watering_data.people))]
        selected = random.choices(watering_data.people, weights=adjusted_weights, k=2)

        # Update selection count and history
        for person in selected:
            selection_count[person] += 1
            watering_data.watering_history[person].append(f"Week {week}")

        schedule.append(f"Week {week}: {selected[0]} and {selected[1]}")
        week += 1

        # Stop at week 52 and create a new table for the next year
        if week > MAX_WEEK:
            save_to_excel(schedule, watering_data.people, watering_data.watering_history)
            schedule = []

    # Save updated watering history to JSON file
    watering_data.save_data()

    return schedule

# Function to display the schedule in the GUI
def show_schedule():
    schedule = generate_schedule()
    result = "\n".join(schedule)
    save_to_excel(
        schedule,
        watering_data.people,
        watering_data.watering_history
    )
    messagebox.showinfo("Gießplan", result)

# Function to generate the schedule in a separate thread
def generate_schedule_threaded():
    def task():
        try:
            # Acquire a lock to prevent race conditions
            threading_lock = threading.Lock()
            with threading_lock:
                schedule = generate_schedule()
                result = "\n".join(schedule)

                # Save to Excel safely
                save_to_excel(
                    schedule,
                    watering_data.people,
                    watering_data.watering_history
                )

                # Update the GUI safely using after()
                root.after(0, lambda: messagebox.showinfo("Gießplan", result))
        except Exception as e:
            # Handle unexpected errors gracefully
            root.after(0, lambda: messagebox.showerror("Error", f"An error occurred: {e}"))

    # Start the thread
    threading.Thread(target=task, daemon=True).start()

# Update current_week during year transition
current_week = datetime.date.today().isocalendar()[1]

# Modify save_to_excel to handle year transitions robustly
def save_to_excel(schedule, people, history, new_year=False):
    try:
        current_year = datetime.date.today().year
        file_name = "Gießplan.xlsx"
        file_exists = os.path.exists(file_name)

        if file_exists:
            try:
                workbook = openpyxl.load_workbook(file_name)
            except Exception as e:
                messagebox.showerror("Error", f"Failed to open Excel file: {e}. Creating a new file.")
                workbook = openpyxl.Workbook()
                file_exists = False

        if not file_exists:
            workbook = openpyxl.Workbook()
            sheet1 = workbook.active
            sheet1.title = "Statistics"
            sheet1.append(["Name", "Watering Count", "Weight"])
            sheet2 = workbook.create_sheet(title=f"Schedule {current_year}")
            sheet2.append(["Week", "Person 1", "Person 2"])

        else:
            if f"Schedule {current_year}" in workbook.sheetnames:
                sheet1 = workbook["Statistics"]
                sheet2 = workbook[f"Schedule {current_year}"]
            else:
                sheet2 = workbook.create_sheet(title=f"Schedule {current_year}")
                sheet2.append(["Week", "Person 1", "Person 2"])

        if new_year:
            next_year = current_year + 1
            if f"Schedule {next_year}" not in workbook.sheetnames:
                sheet2 = workbook.create_sheet(title=f"Schedule {next_year}")
                sheet2.append(["Week", "Person 1", "Person 2"])

            new_json_file = watering_data.yearly_file_template.format(year=next_year)  # Use template for yearly file paths
            new_history = {person: [] for person in people}
            with open(new_json_file, "w") as file:
                json.dump({"PEOPLE": people, "WEIGHTS": watering_data.weights, "WATERING_HISTORY": new_history}, file)

            watering_data.file_path = new_json_file
            watering_data.watering_history = new_history

            # Reset current_week to 1 for the new year
            global current_week
            current_week = 1

        for i, person in enumerate(people):
            watering_count = len(history.get(person, []))
            sheet1.append([person, watering_count, watering_data.weights[i]])

        for i, week in enumerate(schedule):
            try:
                week_number, people = week.split(": ")
                if " and " in people:
                    person1, person2 = people.split(" and ")
                    sheet2.append([week_number, person1, person2])
                else:
                    sheet2.append([week_number, "", ""])
            except ValueError:
                sheet2.append(["", "", ""])

        workbook.save(file_name)
        messagebox.showinfo("Success", f"Plan saved to Excel in sheet 'Schedule {current_year}' of '{file_name}'.")

    except PermissionError:
        messagebox.showerror("Error", "Permission denied while accessing Excel file. Please check file permissions.")
    except Exception as e:
        messagebox.showerror("Error", f"An unexpected error occurred while saving to Excel: {e}. Please contact support.")

def synchronize_watering_history():
    """Synchronize watering_history with the current list of people."""
    people_set = set(watering_data.people)  # Use set for faster lookups

    # Add missing people to watering_history
    for person in people_set:
        if person not in watering_data.watering_history:
            watering_data.watering_history[person] = []

    # Remove entries from watering_history for deleted people
    for person in list(watering_data.watering_history.keys()):
        if person not in people_set:
            del watering_data.watering_history[person]

# Refactor refresh_dependencies to use synchronize_watering_history
def refresh_dependencies():
    synchronize_watering_history()
    update_data()
    watering_data.data_modified = True

# Refactor update_people_list to use synchronize_watering_history
def update_people_list():
    synchronize_watering_history()

    # Update the GUI list
    people_list.delete(0, tk.END)
    for person in watering_data.people:
        people_list.insert(tk.END, person)

# Ensure all dependent data structures and functions are updated after adding or deleting a person

# def refresh_dependencies():
#     # Synchronize watering_history with PEOPLE
#     people_set = set(watering_data.people)  # Use set for faster lookups
#     for person in people_set:
#         if person not in watering_data.watering_history:
#             watering_data.watering_history[person] = []

#     # Remove entries from watering_history for deleted people
#     for person in list(watering_data.watering_history.keys()):
#         if person not in people_set:
#             del watering_data.watering_history[person]

#     # Update weights dynamically
#     update_data()

#     # Mark data as modified
#     watering_data.data_modified = True

# Modify add_person and delete_person to set the data_modified flag
def add_person():
    name = name_entry.get().strip()
    name = sanitize_input(name)  # Sanitize input

    # Validate the name
    if not name.isalpha():
        messagebox.showerror("Error", "Name must contain only letters.")
        return
    if name in watering_data.people:
        messagebox.showerror("Error", "Name already exists.")
        return

    # Push undo action for adding a person
    watering_data.push_undo_action({
        'type': 'add_person',
        'name': name
    })

    # Add the person to the list and update dependencies
    watering_data.people.append(name)
    watering_data.weights.append(1)  # Default weight for new person
    watering_data.watering_history[name] = []
    update_people_list()
    refresh_dependencies()

    # Mark data as modified and show success message
    watering_data.data_modified = True
    messagebox.showinfo("Success", f"Person '{name}' added successfully.")

# Function to delete a person from the list
def delete_person():
    name = name_entry.get()
    name = sanitize_input(name)  # Sanitize input

    # Check if the person exists in the list
    if name in watering_data.people:
        index = watering_data.people.index(name)

        # Push undo action for deleting a person
        watering_data.push_undo_action({
            'type': 'delete_person',
            'name': name,
            'weight': watering_data.weights[index],
            'history': watering_data.watering_history.pop(name, None)
        })

        # Remove the person from the list and update dependencies
        watering_data.people.pop(index)
        watering_data.weights.pop(index)
        update_people_list()
        refresh_dependencies()

        # Mark data as modified
        watering_data.data_modified = True
    else:
        messagebox.showerror("Error", "Name not found.")

# Function to undo the last action
def undo_action():
    action = watering_data.pop_undo_action()

    # Check if there is an action to undo
    if action:
        if action['type'] == 'delete_person':
            # Undo deleting a person
            watering_data.people.append(action['name'])
            watering_data.weights.append(action['weight'])
            watering_data.watering_history[action['name']] = action['history']
            update_people_list()
            refresh_dependencies()
        elif action['type'] == 'delete_date_or_week':
            # Undo deleting a date/week
            for person, entries in action['history'].items():
                watering_data.watering_history[person].extend(entries)
            refresh_dependencies()
        elif action['type'] == 'add_person':
            # Undo adding a person
            name = action['name']
            if name in watering_data.people:
                index = watering_data.people.index(name)
                watering_data.people.pop(index)
                watering_data.weights.pop(index)
                watering_data.watering_history.pop(name, None)
                update_people_list()
                refresh_dependencies()
        elif action['type'] == 'add_date_or_week':
            # Undo adding a date/week
            date_or_week = action['date_or_week']
            person1 = action['person1']
            person2 = action['person2']
            watering_data.watering_history[person1] = [entry for entry in watering_data.watering_history[person1] if not entry.startswith(date_or_week)]
            watering_data.watering_history[person2] = [entry for entry in watering_data.watering_history[person2] if not entry.startswith(date_or_week)]
            refresh_dependencies()

        # Show success message
        messagebox.showinfo("Undo", "Action undone successfully.")
    else:
        messagebox.showerror("Undo", "No actions to undo.")

# Function to delete a specific date or week from the watering schedule
def delete_date_or_week():
    date_or_week = date_entry.get()

    # Validate the input
    if date_or_week:
        # Push undo action for deleting a date/week
        watering_data.push_undo_action({
            'type': 'delete_date_or_week',
            'history': {
                person: [entry for entry in watering_data.watering_history[person] if entry.startswith(date_or_week)]
                for person in watering_data.people
            }
        })

        # Remove the date/week from the watering history
        for person in watering_data.people:
            watering_data.watering_history[person] = [
                entry for entry in watering_data.watering_history[person] if not entry.startswith(date_or_week)
            ]

        # Mark data as modified and save changes
        watering_data.data_modified = True
        save_to_excel([], watering_data.people, watering_data.watering_history)
        messagebox.showinfo("Success", "Date/Week deleted successfully.")
    else:
        messagebox.showerror("Error", "Invalid input. Please provide a date/week.")

# Function to search for people and entries in the watering history
def search_people():
    query = sanitize_input(search_entry.get())

    # Validate the search query
    if not query:
        messagebox.showerror("Error", "Search query cannot be empty.")
        return

    # Filter people based on the query
    filtered_people = [person for person in watering_data.people if query.lower() in person.lower()]

    # Search within watering history
    filtered_history = {
        person: [entry for entry in watering_data.watering_history.get(person, []) if query.lower() in entry.lower()]
        for person in watering_data.people
    }

    # Update the GUI list with filtered results
    people_list.delete(0, tk.END)
    for person in filtered_people:
        people_list.insert(tk.END, person)

    # Display matching history entries in a messagebox
    history_results = "\n".join(
        f"{person}: {', '.join(entries)}" for person, entries in filtered_history.items() if entries
    )
    if history_results:
        messagebox.showinfo("Search Results", history_results)
    else:
        messagebox.showinfo("Search Results", "No matching entries found in watering history.")

# Ensure add_date_or_week is defined before its usage
def add_date_or_week():
    """
    Adds a specific date or week to the watering schedule.
    Validates the input format and updates the watering history and Excel file.
    """
    date_or_week = date_entry.get().strip()
    date_or_week = sanitize_input(date_or_week)  # Sanitize input
    person1 = sanitize_input(person1_entry.get())
    person2 = sanitize_input(person2_entry.get())

    # Validate inputs
    if not date_or_week:
        messagebox.showerror("Error", "Date/Week cannot be empty.")
        return
    if not person1 or not person2:
        messagebox.showerror("Error", "Both Person 1 and Person 2 must be specified.")
        return
    if person1 not in watering_data.people or person2 not in watering_data.people:
        messagebox.showerror("Error", "Both persons must exist in the list.")
        return

    # Validate date/week format
    if date_or_week.startswith("Week"):
        try:
            week_number = int(date_or_week.split()[1])
            if week_number < 1 or week_number > MAX_WEEK:
                raise ValueError("Week number out of range.")
            schedule_entry = f"Week {week_number}: {person1} and {person2}"
        except ValueError:
            messagebox.showerror("Error", "Invalid week format. Use 'Week X' where X is between 1 and 52.")
            return
    else:
        try:
            datetime.datetime.strptime(date_or_week, "%Y-%m-%d")
            schedule_entry = f"Date {date_or_week}: {person1} and {person2}"
        except ValueError:
            messagebox.showerror("Error", "Invalid date format. Use 'YYYY-MM-DD'.")
            return

    # Push undo action for adding a date/week
    watering_data.push_undo_action({
        'type': 'add_date_or_week',
        'date_or_week': date_or_week,
        'person1': person1,
        'person2': person2
    })

    # Update watering history
    watering_data.watering_history[person1].append(schedule_entry)
    watering_data.watering_history[person2].append(schedule_entry)
    watering_data.data_modified = True  # Mark data as modified

    # Save changes and show success message
    save_to_excel([schedule_entry], watering_data.people, watering_data.watering_history)
    messagebox.showinfo("Success", "Date/Week added successfully.")

# Widgets for adding and deleting people
name_label = tk.Label(root, text="Name:")
name_label.pack(pady=5)
name_entry = tk.Entry(root)
name_entry.pack(pady=5)
add_button = tk.Button(root, text="Add Person", command=add_person)
add_button.pack(pady=5)
delete_button = tk.Button(root, text="Delete Person", command=delete_person)
delete_button.pack(pady=5)

# Listbox to display people and their watering history
people_list = tk.Listbox(root, width=50)
people_list.pack(pady=10)
update_people_list()

# Button to generate the schedule
generate_button = tk.Button(root, text="Generate Gießplan", command=generate_schedule_threaded)
generate_button.pack(pady=10)

# Additional GUI elements for adding/deleting specific dates or weeks
label_date = tk.Label(root, text="Date/Week:")
label_date.pack(pady=5)
date_entry = tk.Entry(root)
date_entry.pack(pady=5)

label_person1 = tk.Label(root, text="Person 1:")
label_person1.pack(pady=5)
person1_entry = tk.Entry(root)
person1_entry.pack(pady=5)

label_person2 = tk.Label(root, text="Person 2:")
label_person2.pack(pady=5)
person2_entry = tk.Entry(root)
person2_entry.pack(pady=5)

add_date_button = tk.Button(root, text="Add Date/Week", command=add_date_or_week)
add_date_button.pack(pady=5)
delete_date_button = tk.Button(root, text="Delete Date/Week", command=delete_date_or_week)
delete_date_button.pack(pady=5)

# Add search/filter functionality GUI elements
search_label = tk.Label(root, text="Search:")
search_label.pack(pady=5)
search_entry = tk.Entry(root)
search_entry.pack(pady=5)
search_button = tk.Button(root, text="Search", command=search_people)
search_button.pack(pady=5)

root.mainloop()
