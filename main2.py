import tkinter as tk
from tkinter import messagebox
import random
import openpyxl
import json
import os
import datetime

# Define the file path for storing names
FILE_PATH = "people.json"

# Load names from file or use default ones
if os.path.exists(FILE_PATH):
    try:
        with open(FILE_PATH, "r") as file:
            data = json.load(file)
            PEOPLE = data.get("PEOPLE", [])
            WEIGHTS = data.get("WEIGHTS", [])
            watering_history = data.get("WATERING_HISTORY", {})
    except json.JSONDecodeError:
        # Handle empty or invalid JSON file
        PEOPLE = ["Alice", "Bob", "Charlie", "Diana", "Eve", "Frank"]
        WEIGHTS = [1, 2, 1, 1, 2, 1]
        watering_history = {person: [] for person in PEOPLE}
        with open(FILE_PATH, "w") as file:
            json.dump({"PEOPLE": PEOPLE, "WEIGHTS": WEIGHTS, "WATERING_HISTORY": watering_history}, file)
else:
    PEOPLE = ["Alice", "Bob", "Charlie", "Diana", "Eve", "Frank"]
    WEIGHTS = [1, 2, 1, 1, 2, 1]  # Adjust weights as needed
    watering_history = {person: [] for person in PEOPLE}
    with open(FILE_PATH, "w") as file:
        json.dump({"PEOPLE": PEOPLE, "WEIGHTS": WEIGHTS, "WATERING_HISTORY": watering_history}, file)

# Ensure names in PEOPLE contain only letters
PEOPLE = [person for person in PEOPLE if person.isalpha()]

# Function to generate the watering schedule with interval-based weight adjustment
# Start with the current calendar week
current_week = datetime.date.today().isocalendar()[1]

# Function to find the next free calendar week
def find_next_free_week():
    used_weeks = set()
    for person, history in watering_history.items():
        if isinstance(history, list):  # Ensure history is iterable
            for entry in history:
                if entry.startswith("Week"):
                    week_number = int(entry.split()[1])
                    used_weeks.add(week_number)

    next_week = current_week
    while next_week in used_weeks:
        next_week += 1

    return next_week

# Update watering counts dynamically based on watering history
def update_statistics():
    for person in PEOPLE:
        # Calculate watering count based on history
        watering_count = len(watering_history.get(person, []))
        WEIGHTS[PEOPLE.index(person)] = max(1, 10 - watering_count)  # Example: higher weight for fewer waterings

    # Save updated statistics to JSON file
    save_to_file()

# Update weights dynamically based on watering history
def update_weights():
    for person in PEOPLE:
        # Calculate weight based on the number of times the person has watered
        watering_count = len(watering_history.get(person, []))
        WEIGHTS[PEOPLE.index(person)] = max(1, 10 - watering_count)  # Example: higher weight for fewer waterings

    # Save updated weights to JSON file
    save_to_file()

def generate_schedule():
    update_statistics()  # Update statistics before generating the schedule

    schedule = []
    selection_count = {person: len(watering_history[person]) for person in PEOPLE}  # Track how often each person is selected

    # Get the current year
    current_year = datetime.date.today().year

    # Determine the starting week based on the last recorded week in the JSON file
    if any(entry.startswith("Week") for person in watering_history.values() for entry in person):
        last_week = max(
            [int(entry.split()[1]) for person in watering_history.values() for entry in person if entry.startswith("Week")]
        )
        start_week = last_week + 1
    else:
        start_week = current_week + 1

    max_week = 52  # Define the maximum number of weeks in a year

    week = start_week
    while len(schedule) < 6:  # Generate 6 weeks dynamically
        if week > max_week:
            # Transition to the next year
            current_year += 1
            week = 1
            watering_history["last_year"] = current_year

            # Create a new table in the Excel file for the new year
            save_to_excel([], PEOPLE, watering_history, new_year=True)

        # Adjust weights based on selection intervals
        adjusted_weights = [WEIGHTS[i] / (1 + selection_count[PEOPLE[i]]) for i in range(len(PEOPLE))]
        selected = random.choices(PEOPLE, weights=adjusted_weights, k=2)

        # Update selection count and history
        for person in selected:
            selection_count[person] += 1
            watering_history[person].append(f"Week {week}")

        schedule.append(f"Week {week}: {selected[0]} and {selected[1]}")
        week += 1

        # Stop at week 52 and create a new table for the next year
        if week > max_week:
            save_to_excel(schedule, PEOPLE, watering_history)
            schedule = []

    # Save updated watering history to JSON file
    save_to_file()

    return schedule

# Function to display the schedule in the GUI
def show_schedule():
    schedule = generate_schedule()
    result = "\n".join(schedule)
    save_to_excel(
        schedule,
        PEOPLE,
        watering_history
    )
    messagebox.showinfo("Gießplan", result)

# Modify save_to_excel to handle cases where schedule entries are not properly formatted

def save_to_excel(schedule, people, history, new_year=False):
    # Get the current year
    current_year = datetime.date.today().year

    # Determine the file name
    file_name = "Gießplan.xlsx"

    # Check if the file already exists
    file_exists = os.path.exists(file_name)

    if file_exists:
        workbook = openpyxl.load_workbook(file_name)
        # Check if the sheet for the current year exists
        if f"Schedule {current_year}" in workbook.sheetnames:
            sheet1 = workbook["Statistics"]
            sheet2 = workbook[f"Schedule {current_year}"]
        else:
            # Create a new sheet for the current year
            sheet2 = workbook.create_sheet(title=f"Schedule {current_year}")
            sheet2.append(["Week", "Person 1", "Person 2"])

        # Handle new year transition
        if new_year:
            next_year = current_year + 1
            if f"Schedule {next_year}" not in workbook.sheetnames:
                sheet2 = workbook.create_sheet(title=f"Schedule {next_year}")
                sheet2.append(["Week", "Person 1", "Person 2"])

            # Create a new JSON file for the new year
            new_json_file = f"people_{next_year}.json"
            new_history = {person: [] for person in people}
            with open(new_json_file, "w") as file:
                json.dump({"PEOPLE": people, "WEIGHTS": WEIGHTS, "WATERING_HISTORY": new_history}, file)

            # Update the global FILE_PATH to point to the new JSON file
            global FILE_PATH
            FILE_PATH = new_json_file

            # Update the history reference to the new history
            global watering_history
            watering_history = new_history

    else:
        workbook = openpyxl.Workbook()

        # Create the statistics sheet
        sheet1 = workbook.active
        sheet1.title = "Statistics"
        sheet1.append(["Name", "Watering Count", "Weight"])

        # Create the schedule sheet for the current year
        sheet2 = workbook.create_sheet(title=f"Schedule {current_year}")
        sheet2.append(["Week", "Person 1", "Person 2"])

    # Update statistics sheet with refreshed data
    for i, person in enumerate(people):
        watering_count = len(history.get(person, []))
        sheet1.append([person, watering_count, WEIGHTS[i]])

    # Append the new schedule below the existing rows in the schedule sheet
    for i, week in enumerate(schedule):
        try:
            week_number, people = week.split(": ")
            if " and " in people:
                person1, person2 = people.split(" and ")
                if int(week_number.split()[1]) > 52 and new_year:
                    next_year = current_year + 1
                    sheet2 = workbook[f"Schedule {next_year}"]
                sheet2.append([week_number, person1, person2])
            else:
                sheet2.append([week_number, "", ""])
        except ValueError:
            sheet2.append(["", "", ""])

    # Save the workbook
    workbook.save(file_name)
    messagebox.showinfo("Success", f"Plan saved to Excel in sheet 'Schedule {current_year}' of '{file_name}'.")

# Ensure all dependent data structures and functions are updated after adding or deleting a person

def refresh_dependencies():
    # Synchronize watering_history with PEOPLE
    for person in PEOPLE:
        if person not in watering_history:
            watering_history[person] = []

    # Remove entries from watering_history for deleted people
    for person in list(watering_history.keys()):
        if person not in PEOPLE:
            del watering_history[person]

    # Update weights dynamically
    update_weights()

    # Save changes to JSON file
    save_to_file()

# Modify add_person and delete_person to refresh dependencies

def add_person():
    name = name_entry.get()
    if name.isalpha() and name not in PEOPLE:
        PEOPLE.append(name)
        WEIGHTS.append(1)  # Default weight for new person
        watering_history[name] = []
        update_people_list()
        refresh_dependencies()
    else:
        messagebox.showerror("Error", "Invalid or duplicate name.")

def delete_person():
    name = name_entry.get()
    if name in PEOPLE:
        index = PEOPLE.index(name)
        PEOPLE.pop(index)
        WEIGHTS.pop(index)
        watering_history.pop(name, None)
        update_people_list()
        refresh_dependencies()
    else:
        messagebox.showerror("Error", "Name not found.")

# Update the JSON file to track added and deleted dates/weeks
def save_to_file():
    with open(FILE_PATH, "w") as file:
        json.dump({"PEOPLE": PEOPLE, "WEIGHTS": WEIGHTS, "WATERING_HISTORY": watering_history}, file)

# Function to add a specific date or week to the schedule
def add_date_or_week():
    date_or_week = date_entry.get()
    person1 = person1_entry.get()
    person2 = person2_entry.get()

    if date_or_week and person1 and person2:
        if date_or_week.startswith("Week"):
            week_number = date_or_week.split()[1]
            schedule_entry = f"Week {week_number}: {person1} and {person2}"
        else:
            schedule_entry = f"Date {date_or_week}: {person1} and {person2}"

        # Append to watering history
        watering_history[person1].append(schedule_entry)
        watering_history[person2].append(schedule_entry)

        # Save changes to JSON file
        save_to_file()

        # Save to Excel
        save_to_excel([schedule_entry], PEOPLE, watering_history)
        messagebox.showinfo("Success", "Date/Week added successfully.")
    else:
        messagebox.showerror("Error", "Invalid input. Please provide a date/week and two people.")

# Function to delete a specific date or week from the schedule
def delete_date_or_week():
    date_or_week = date_entry.get()

    if date_or_week:
        for person in PEOPLE:
            watering_history[person] = [entry for entry in watering_history[person] if not entry.startswith(date_or_week)]

        # Save changes to JSON file
        save_to_file()

        # Save updated history to Excel
        save_to_excel([], PEOPLE, watering_history)
        messagebox.showinfo("Success", "Date/Week deleted successfully.")
    else:
        messagebox.showerror("Error", "Invalid input. Please provide a date/week.")

# Ensure watering_history is synchronized with PEOPLE
# Function to update the list of people in the GUI
def update_people_list():
    # Synchronize watering_history with valid PEOPLE
    for person in PEOPLE:
        if person not in watering_history:
            watering_history[person] = []

    # Remove entries from watering_history for deleted people
    for person in list(watering_history.keys()):
        if person not in PEOPLE:
            del watering_history[person]

    # Update the GUI list
    people_list.delete(0, tk.END)
    for person in PEOPLE:
        people_list.insert(tk.END, person)

# Create the GUI
root = tk.Tk()
root.title("Gießplan Generator")

# Widgets for adding and deleting people
name_label = tk.Label(root, text="Name:")
name_label.pack(pady=5)
name_entry = tk.Entry(root)
name_entry.pack(pady=5)
add_button = tk.Button(root, text="Add Person", command=add_person)
add_button.pack(pady=5)
delete_button = tk.Button(root, text="Delete Person", command=delete_person)
delete_button.pack(pady=5)

# Listbox to display people and their watering history
people_list = tk.Listbox(root, width=50)
people_list.pack(pady=10)
update_people_list()

# Button to generate the schedule
generate_button = tk.Button(root, text="Generate Gießplan", command=show_schedule)
generate_button.pack(pady=10)

# Additional GUI elements for adding/deleting specific dates or weeks
label_date = tk.Label(root, text="Date/Week:")
label_date.pack(pady=5)
date_entry = tk.Entry(root)
date_entry.pack(pady=5)

label_person1 = tk.Label(root, text="Person 1:")
label_person1.pack(pady=5)
person1_entry = tk.Entry(root)
person1_entry.pack(pady=5)

label_person2 = tk.Label(root, text="Person 2:")
label_person2.pack(pady=5)
person2_entry = tk.Entry(root)
person2_entry.pack(pady=5)

add_date_button = tk.Button(root, text="Add Date/Week", command=add_date_or_week)
add_date_button.pack(pady=5)
delete_date_button = tk.Button(root, text="Delete Date/Week", command=delete_date_or_week)
delete_date_button.pack(pady=5)

root.mainloop()
