import json
import os
import openpyxl
import datetime
from tkinter import messagebox

FILE_PATH = "people.json"

# Function to migrate data from people.json to people_YEAR.json format
def migrate_people_json_to_year_specific():
    """Migrate data from people.json to people_YEAR.json format"""
    current_year = datetime.date.today().year
    
    # If people.json exists and people_YEAR.json doesn't exist, migrate it
    if os.path.exists("people.json") and not os.path.exists(f"people_{current_year}.json"):
        try:
            # Read the current people.json
            with open("people.json", "r") as file:
                data = json.load(file)
            
            # Save it as people_YEAR.json
            with open(f"people_{current_year}.json", "w") as file:
                json.dump(data, file, indent=2)
            
            print(f"Migrated people.json to people_{current_year}.json")
            return True
        except Exception as e:
            print(f"Error migrating people.json: {e}")
            return False
    return False

def ensure_year_specific_files():
    """Ensure all data is in year-specific file format"""
    current_year = datetime.date.today().year
    
    # Migrate people.json to year-specific format
    if migrate_people_json_to_year_specific():
        print(f"Successfully migrated people.json to people_{current_year}.json")
    
    # Update FILE_PATH to use current year file
    global FILE_PATH
    FILE_PATH = f"people_{current_year}.json"

# Function to get all available year files
def get_available_years():
    """Get all available years from existing data files"""
    years = []
    current_year = datetime.date.today().year
    
    # Check for default file (current year)
    if os.path.exists("people.json"):
        years.append(current_year)
    
    # Check for year-specific files
    for year in range(current_year - 5, current_year + 10):  # Check wide range
        year_file = f"people_{year}.json"
        if os.path.exists(year_file):
            years.append(year)
    
    # Remove duplicates and sort
    years = sorted(list(set(years)))
    return years

# Function to load data from specific year
def load_year_data(year):
    """Load data from a specific year file"""
    global FILE_PATH, PEOPLE, WEIGHTS, watering_history
    
    current_year = datetime.date.today().year
    
    # First, migrate people.json to year-specific format if needed
    if year == current_year:
        migrate_people_json_to_year_specific()
    
    # Always use year-specific file format
    target_file = f"people_{year}.json"
    
    # Try to load the file
    if os.path.exists(target_file):
        try:
            with open(target_file, "r") as file:
                data = json.load(file)
                PEOPLE.clear()
                PEOPLE.extend(data.get("PEOPLE", []))
                WEIGHTS.clear()
                WEIGHTS.extend(data.get("WEIGHTS", []))
                watering_history.clear()
                watering_history.update(data.get("WATERING_HISTORY", {}))
                FILE_PATH = target_file
                return True
        except json.JSONDecodeError:
            return False
    else:
        # File doesn't exist, create it only if switching to current year
        if year == current_year:
            # Create default file for current year
            PEOPLE.clear()
            PEOPLE.extend(["Jan", "Jeff", "Antonia", "Melissa", "Rosa", "Alexander"])
            WEIGHTS.clear()
            WEIGHTS.extend([1, 2, 1, 1, 2, 1])
            watering_history.clear()
            watering_history.update({person: [] for person in PEOPLE})
            FILE_PATH = target_file
            
            try:
                with open(target_file, "w") as file:
                    json.dump({"PEOPLE": PEOPLE, "WEIGHTS": WEIGHTS, "WATERING_HISTORY": watering_history}, file, indent=2)
                return True
            except PermissionError:
                print(f"Permission error writing to {target_file} - file may be open in another application")
                return False
            except Exception as e:
                print(f"Error writing to {target_file}: {str(e)}")
                return False
        else:
            # Don't create files for other years automatically
            return False

# Function to get the current year's people file
def get_current_people_file():
    current_year = datetime.date.today().year
    
    # First, try to get the current year's file
    current_year_file = f"people_{current_year}.json"
    if os.path.exists(current_year_file):
        return current_year_file
    
    # If current year file doesn't exist, check for people.json
    if os.path.exists("people.json"):
        return "people.json"
    
    # If neither exists, return the current year file path (will be created)
    return current_year_file

# Migrate people.json to year-specific format on startup
current_year = datetime.date.today().year
migrate_people_json_to_year_specific()

# Always start with the current year (2025)
FILE_PATH = f"people_{current_year}.json"

# Load names from file or use default ones
if os.path.exists(FILE_PATH):
    try:
        with open(FILE_PATH, "r") as file:
            data = json.load(file)
            PEOPLE = data.get("PEOPLE", [])
            WEIGHTS = data.get("WEIGHTS", [])
            watering_history = data.get("WATERING_HISTORY", {})
    except json.JSONDecodeError:
        PEOPLE = ["Alice", "Bob", "Charlie", "Diana", "Eve", "Frank"]
        WEIGHTS = [1, 2, 1, 1, 2, 1]
        watering_history = {person: [] for person in PEOPLE}
        with open(FILE_PATH, "w") as file:
            json.dump({"PEOPLE": PEOPLE, "WEIGHTS": WEIGHTS, "WATERING_HISTORY": watering_history}, file)
else:
    PEOPLE = ["Jan", "Jeff", "Antonia", "Melissa", "Rosa", "Alexander"]
    WEIGHTS = [1, 2, 1, 1, 2, 1]
    watering_history = {person: [] for person in PEOPLE}
    with open(FILE_PATH, "w") as file:
        json.dump({"PEOPLE": PEOPLE, "WEIGHTS": WEIGHTS, "WATERING_HISTORY": watering_history}, file)

def get_current_year():
    """Get the year currently being worked on"""
    if FILE_PATH.endswith("people.json"):
        return datetime.date.today().year
    else:
        # Extract year from filename like "people_2026.json"
        import re
        match = re.search(r'people_(\d{4})\.json', FILE_PATH)
        return int(match.group(1)) if match else datetime.date.today().year

def reload_current_data():
    """Reload data from the currently selected file"""
    global FILE_PATH, PEOPLE, WEIGHTS, watering_history
    
    # Use the current FILE_PATH instead of getting the most recent file
    if os.path.exists(FILE_PATH):
        try:
            with open(FILE_PATH, "r") as file:
                data = json.load(file)
                PEOPLE.clear()
                PEOPLE.extend(data.get("PEOPLE", []))
                WEIGHTS.clear()
                WEIGHTS.extend(data.get("WEIGHTS", []))
                watering_history.clear()
                watering_history.update(data.get("WATERING_HISTORY", {}))
        except json.JSONDecodeError:
            pass  # Keep current data if file is corrupted

def save_to_file():
    global FILE_PATH
    # Don't change FILE_PATH here - use the current one
    try:
        with open(FILE_PATH, "w") as file:
            json.dump({"PEOPLE": PEOPLE, "WEIGHTS": WEIGHTS, "WATERING_HISTORY": watering_history}, file, indent=2)
    except PermissionError:
        print(f"Permission error writing to {FILE_PATH} - file may be open in another application")
        raise PermissionError(f"Cannot write to {FILE_PATH} - file may be open in another application")
    except Exception as e:
        print(f"Error writing to {FILE_PATH}: {str(e)}")
        raise

def update_weights():
    """Update weights based on watering history to maintain balance"""
    # Calculate total weeks active in system - filter out non-list entries
    all_week_entries = []
    for entries in watering_history.values():
        if isinstance(entries, list):
            for entry in entries:
                if entry.startswith("Week"):
                    all_week_entries.append(entry)
    
    total_weeks_active = len(set(all_week_entries)) if all_week_entries else 1
    
    for i, person in enumerate(PEOPLE):
        watering_count = len(watering_history.get(person, []))
        
        # Adaptive weight calculation based on system duration
        if total_weeks_active <= 4:
            # Short-term: Higher weights for new people
            WEIGHTS[i] = max(1, 8 - watering_count)
        elif total_weeks_active <= 52:
            # Medium-term: Balanced approach
            expected_waterings = total_weeks_active * 2 / len(PEOPLE)
            WEIGHTS[i] = max(1, int(expected_waterings - watering_count + 5))
        else:
            # Long-term: Focus on recent balance, normalized weights
            expected_waterings = total_weeks_active * 2 / len(PEOPLE)
            deviation = expected_waterings - watering_count
            WEIGHTS[i] = max(1, int(5 + deviation))
    
    save_to_file()

def save_to_excel(schedule, people, history, new_year=False, target_year=None):
    file_name = "Gießplan.xlsx"
    current_year = datetime.date.today().year
    
    # Determine which year we're working with
    if target_year:
        schedule_year = target_year
    elif new_year:
        schedule_year = current_year + 1
    else:
        schedule_year = current_year
    
    file_exists = os.path.exists(file_name)

    if file_exists:
        workbook = openpyxl.load_workbook(file_name)
        
        # Ensure Statistics sheet exists
        if "Statistics" not in workbook.sheetnames:
            sheet1 = workbook.create_sheet(title="Statistics")
            sheet1.append(["Name", "Watering Count", "Weight"])
        else:
            sheet1 = workbook["Statistics"]
            # Clear existing statistics
            sheet1.delete_rows(2, sheet1.max_row)
        
        # Ensure target year schedule sheet exists
        if f"Schedule {schedule_year}" not in workbook.sheetnames:
            sheet2 = workbook.create_sheet(title=f"Schedule {schedule_year}")
            sheet2.append(["Week", "Person 1", "Person 2"])
        else:
            sheet2 = workbook[f"Schedule {schedule_year}"]
            # Always clear existing schedule entries when saving new data
            # This prevents accumulation of old entries
            if sheet2.max_row > 1:
                sheet2.delete_rows(2, sheet2.max_row)

        if new_year:
            # When creating a new year, create the new JSON file
            new_json_file = f"people_{schedule_year}.json"
            new_history = {person: [] for person in people}
            with open(new_json_file, "w") as file:
                json.dump({"PEOPLE": people, "WEIGHTS": WEIGHTS, "WATERING_HISTORY": new_history}, file)

    else:
        workbook = openpyxl.Workbook()
        sheet1 = workbook.active
        sheet1.title = "Statistics"
        sheet1.append(["Name", "Watering Count", "Weight"])
        sheet2 = workbook.create_sheet(title=f"Schedule {schedule_year}")
        sheet2.append(["Week", "Person 1", "Person 2"])

    # Update statistics
    for i, person in enumerate(people):
        watering_count = len(history.get(person, []))
        if i < len(WEIGHTS):
            sheet1.append([person, watering_count, WEIGHTS[i]])
        else:
            sheet1.append([person, watering_count, 1])

    # Add schedule entries to the correct year's sheet
    for i, week in enumerate(schedule):
        try:
            week_number, people_str = week.split(": ")
            if " and " in people_str:
                person1, person2 = people_str.split(" and ")
                sheet2.append([week_number, person1, person2])
            else:
                sheet2.append([week_number, "", ""])
        except ValueError:
            sheet2.append(["", "", ""])

    workbook.save(file_name)
    if schedule:  # Only show success message if there are actual schedule entries
        messagebox.showinfo("Success", f"Plan saved to Excel in sheet 'Schedule {schedule_year}' of '{file_name}'.")

def save_new_weeks_to_excel(new_weeks, people, history, target_year=None):
    """Save only new weeks to Excel, appending to existing schedule"""
    file_name = "Gießplan.xlsx"
    current_year = datetime.date.today().year
    
    # Determine which year we're working with
    if target_year:
        schedule_year = target_year
    else:
        schedule_year = current_year
    
    file_exists = os.path.exists(file_name)

    if file_exists:
        workbook = openpyxl.load_workbook(file_name)
        
        # Ensure Statistics sheet exists
        if "Statistics" not in workbook.sheetnames:
            sheet1 = workbook.create_sheet(title="Statistics")
            sheet1.append(["Name", "Watering Count", "Weight"])
        else:
            sheet1 = workbook["Statistics"]
            # Clear existing statistics
            sheet1.delete_rows(2, sheet1.max_row)
        
        # Ensure target year schedule sheet exists
        if f"Schedule {schedule_year}" not in workbook.sheetnames:
            sheet2 = workbook.create_sheet(title=f"Schedule {schedule_year}")
            sheet2.append(["Week", "Person 1", "Person 2"])
        else:
            sheet2 = workbook[f"Schedule {schedule_year}"]
            # DO NOT clear existing schedule entries - we want to append

    else:
        workbook = openpyxl.Workbook()
        sheet1 = workbook.active
        sheet1.title = "Statistics"
        sheet1.append(["Name", "Watering Count", "Weight"])
        sheet2 = workbook.create_sheet(title=f"Schedule {schedule_year}")
        sheet2.append(["Week", "Person 1", "Person 2"])

    # Update statistics
    for i, person in enumerate(people):
        watering_count = len(history.get(person, []))
        if i < len(WEIGHTS):
            sheet1.append([person, watering_count, WEIGHTS[i]])
        else:
            sheet1.append([person, watering_count, 1])

    # Add only the new schedule entries to the correct year's sheet
    for i, week in enumerate(new_weeks):
        try:
            week_number, people_str = week.split(": ")
            if " and " in people_str:
                person1, person2 = people_str.split(" and ")
                sheet2.append([week_number, person1, person2])
            else:
                sheet2.append([week_number, "", ""])
        except ValueError:
            sheet2.append(["", "", ""])

    workbook.save(file_name)
    if new_weeks:  # Only show success message if there are actual new schedule entries
        messagebox.showinfo("Success", f"Plan saved to Excel in sheet 'Schedule {schedule_year}' of '{file_name}'.")

def refresh_dependencies():

    for person in PEOPLE:
        if person not in watering_history:
            watering_history[person] = []
    for person in list(watering_history.keys()):
        if person not in PEOPLE:
            del watering_history[person]
    update_weights()
    save_to_file()

def add_new_person_with_context(name, join_week=None):
    """Add a new person with weight equal to the average of all other people combined"""
    if name in PEOPLE:
        return False
    
    # Calculate the average weight of all existing people
    if PEOPLE and WEIGHTS:
        # Calculate average weight of all existing people
        average_weight = sum(WEIGHTS) / len(WEIGHTS)
        # Set new person's weight to the average of all others combined
        initial_weight = max(1, int(round(average_weight)))
    else:
        # If no people exist yet, use default weight
        initial_weight = 5
    
    PEOPLE.append(name)
    WEIGHTS.append(initial_weight)
    watering_history[name] = []
    save_to_file()
    return True

def remove_person_and_rebalance(name):
    """Remove a person and rebalance the system"""
    if name not in PEOPLE:
        return False
    
    # Store the person's watering history before removal
    leaving_person_waterings = len(watering_history.get(name, []))
    
    # Remove the person
    index = PEOPLE.index(name)
    PEOPLE.pop(index)
    WEIGHTS.pop(index)
    watering_history.pop(name, None)
    
    # Rebalance remaining people's weights
    # If someone was carrying more load, adjust others accordingly
    if leaving_person_waterings > 0 and len(PEOPLE) > 0:
        # Distribute some of the leaving person's "load" to others
        load_redistribution = max(1, leaving_person_waterings // len(PEOPLE))
        for i in range(len(PEOPLE)):
            WEIGHTS[i] = max(1, WEIGHTS[i] - load_redistribution)
    
    update_weights()
    save_to_file()
    return True

def get_week_data(year, week):
    """Find the two people assigned for a given year and week."""
    target_file = f"people_{year}.json"
    if os.path.exists(target_file):
        with open(target_file, "r") as file:
            data = json.load(file)
        wh = data.get("WATERING_HISTORY", {})
        week_str = f"{year} KW {week}:"
        for entries in wh.values():
            for entry in entries:
                if entry.startswith(week_str):
                    # Example entry: "2025 KW 30: Jan and Jeff"
                    try:
                        people_part = entry.split(":")[1].strip()
                        person1, _, person2 = people_part.partition(" and ")
                        return [person1.strip(), person2.strip()]
                    except Exception:
                        continue
    return ["", ""]

def update_week_data(year, week, person1, person2):
    """Update data for a specific week in a given year."""
    global watering_history, FILE_PATH
    
    target_file = f"people_{year}.json"
    if os.path.exists(target_file):
        with open(target_file, "r") as file:
            data = json.load(file)
    else:
        # Create a new file with the expected structure
        data = {
            "PEOPLE": PEOPLE.copy(),
            "WEIGHTS": WEIGHTS.copy(),
            "WATERING_HISTORY": {person: [] for person in PEOPLE}
        }

    # Get the watering history
    watering_history_data = data.get("WATERING_HISTORY", {})
    
    # Create the week entry
    week_entry = f"{year} KW {week}: {person1} and {person2}"
    
    # Remove any existing entries for this week
    week_str = f"{year} KW {week}:"
    for person in watering_history_data:
        if isinstance(watering_history_data[person], list):
            watering_history_data[person] = [entry for entry in watering_history_data[person] if not entry.startswith(week_str)]
    
    # Add the new entry to both people's history
    if person1 not in watering_history_data:
        watering_history_data[person1] = []
    if person2 not in watering_history_data:
        watering_history_data[person2] = []
    
    watering_history_data[person1].append(week_entry)
    watering_history_data[person2].append(week_entry)
    
    # Update the data structure
    data["WATERING_HISTORY"] = watering_history_data

    # Save back to the file
    with open(target_file, "w") as file:
        json.dump(data, file, indent=2)
    
    # If we're updating the current file, also update global variables
    if target_file == FILE_PATH:
        watering_history.clear()
        watering_history.update(watering_history_data)


